import tkinter
import openpyxl
from config.definitions import ROOT_DIR
import os

wb = load_workbook(os.path.join(ROOT_DIR,"resources/0.xls"))

sheet = wb.active


def excel():
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	sheet.cell(row=1, column=1).value = "Denumire"
	sheet.cell(row=1, column=2).value = "Lot"
	sheet.cell(row=1, column=3).value = "Concentratie"
	sheet.cell(row=1, column=4).value = "Bbd"
	sheet.cell(row=1, column=5).value = "Serie"
	sheet.cell(row=1, column=6).value = "Email pr"
	sheet.cell(row=1, column=7).value = "Adresa"

def focus1(event):
	lot_field.focus_set()

def focus2(event):
	concentratie_field.focus_set()

def focus3(event):
	bbd_field.focus_set()

def focus4(event):
	serie_field.focus_set()

def focus5(event):
	email_pr_field.focus_set()

def focus6(event):
	adresa_field.focus_set()

def clear():
	denumire_field.delete(0, END)
	lot_field.delete(0, END)
	concentratie_field.delete(0, END)
	bbd_field.delete(0, END)
	serie_field.delete(0, END)
	email_pr_field.delete(0, END)
	adresa_field.delete(0, END)

def insert():
	if (denumire_field.get() == "" and
		lot_field.get() == "" and
		concentratie_field.get() == "" and
		bbd_field.get() == "" and
		serie_field.get() == "" and
		email_pr_field.get() == "" and
		adresa_field.get() == ""):
			
		print("empty input")

	else:
		current_row = sheet.max_row
		current_column = sheet.max_column
        
		sheet.cell(row=current_row + 1, column=1).value = denumire_field.get()
		sheet.cell(row=current_row + 1, column=2).value = lot_field.get()
		sheet.cell(row=current_row + 1, column=3).value = concentratie_field.get()
		sheet.cell(row=current_row + 1, column=4).value = bbd_field.get()
		sheet.cell(row=current_row + 1, column=5).value = serie_field.get()
		sheet.cell(row=current_row + 1, column=6).value = email_pr_field.get()
		sheet.cell(row=current_row + 1, column=7).value = adresa_field.get()

		wb.save(os.path.join(ROOT_DIR,"output/0.xls"))

		denumire_field.focus_set()

		clear()


if __denumire__ == "__main__":
	
	root = Tk()

	root.configure(background='light green')

	root.title("proces verbal")

	root.geometry("500x300")

	excel()

	heading = Label(root, text="Form", bg="light green")

	denumire = Label(root, text="Denumire", bg="light green")

	lot = Label(root, text="Lot", bg="light green")

	concentratie = Label(root, text="Concentratie", bg="light green")

	bbd = Label(root, text="Bbd", bg="light green")

	serie = Label(root, text="Serie", bg="light green")

	email_pr = Label(root, text="Email pr", bg="light green")

	adresa = Label(root, text="Adresa", bg="light green")


	heading.grid(row=0, column=1)
	denumire.grid(row=1, column=0)
	lot.grid(row=2, column=0)
	concentratie.grid(row=3, column=0)
	bbd.grid(row=4, column=0)
	serie.grid(row=5, column=0)
	email_pr.grid(row=6, column=0)
	adresa.grid(row=7, column=0)


	denumire_field = Entry(root)
	lot_field = Entry(root)
	concentratie_field = Entry(root)
	bbd_field = Entry(root)
	serie_field = Entry(root)
	email_pr_field = Entry(root)
	adresa_field = Entry(root)


	denumire_field.bind("<Return>", focus1)

	lot_field.bind("<Return>", focus2)

	concentratie_field.bind("<Return>", focus3)

	bbd_no_field.bind("<Return>", focus4)

	serie_field.bind("<Return>", focus5)

	email_pr_field.bind("<Return>", focus6)

	denumire_field.grid(row=1, column=1, ipadx="100")
	lot_field.grid(row=2, column=1, ipadx="100")
	concentratie_field.grid(row=3, column=1, ipadx="100")
	bbd_field.grid(row=4, column=1, ipadx="100")
	serie_field.grid(row=5, column=1, ipadx="100")
	email_pr_field.grid(row=6, column=1, ipadx="100")
	adresa_field.grid(row=7, column=1, ipadx="100")


	excel()


	submit = Button(root, text="Submit", fg="Black",
							bg="Red", command=insert)
	submit.grid(row=8, column=1)

	root.mainloop()
